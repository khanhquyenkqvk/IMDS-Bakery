# backend/views/owner_report_views.py
from flask import Blueprint, request, jsonify
from datetime import date, datetime, timedelta
from decimal import Decimal
from backend.utils.db import get_conn, dictfetchall
from flask import send_file
import io
from openpyxl import Workbook


bp_owner_reports = Blueprint('bp_owner_reports', __name__)

# ===== Helpers =====

def _parse_date(s, default):
    if not s:
        return default
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return default

def _get_range():
    """
    Lấy khoảng ngày [start, end] từ query string.
    Mặc định: 30 ngày gần nhất.
    end_date là inclusive → khi query dùng BETWEEN start AND end+1.
    """
    today = date.today()
    end = _parse_date(request.args.get("end_date"), today)
    start = _parse_date(request.args.get("start_date"), end - timedelta(days=30))
    return start, end

def _to_float(x):
    if x is None:
        return 0.0
    if isinstance(x, Decimal):
        return float(x)
    return float(x)

def _unit_to_kg(qty, unit):
    """Convert quantity + unit → kg tương đối (đủ dùng cho report)."""
    if qty is None:
        return 0.0
    q = _to_float(qty)
    u = (unit or "").lower()
    if u in ("kg", "kilogram", "kilograms"):
        return q
    if u in ("g", "gram", "grams"):
        return q / 1000.0
    if u in ("mg",):
        return q / 1_000_000.0
    if u in ("l", "liter", "litre", "liters", "ml"):
        # tạm coi 1L ~ 1kg cho nguyên liệu dạng nước/sữa
        if u == "ml":
            return q / 1000.0
        return q
    # đơn vị "pcs", "units" → không quy đổi, dùng luôn
    return q
@bp_owner_reports.route("/api/owner/reports/overview", methods=["GET"])
def overview():
    start, end = _get_range()
    end_inclusive = end + timedelta(days=1)

    conn = get_conn()
    cur = conn.cursor()
    try:
        # Restock events (Import)
        cur.execute("""
            SELECT COUNT(*) 
            FROM transactions
            WHERE type = 'Import'
              AND created_at BETWEEN %s AND %s
        """, (start, end_inclusive))
        total_restock = cur.fetchone()[0] or 0

        # Active ingredients (đang còn tồn)
        cur.execute("""
            SELECT COUNT(*)
            FROM inventory
            WHERE current_stock > 0
        """)
        active_items = cur.fetchone()[0] or 0

        # Waste rate = Waste / (Use + Waste + Export)
        cur.execute("""
            SELECT
              SUM(CASE WHEN type='Waste' THEN quantity ELSE 0 END) AS waste_qty,
              SUM(CASE WHEN type IN ('Use','Waste','Export') THEN quantity ELSE 0 END) AS total_out
            FROM transactions
            WHERE created_at BETWEEN %s AND %s
        """, (start, end_inclusive))
        waste_qty, total_out = cur.fetchone()
        waste_qty = _to_float(waste_qty)
        total_out = _to_float(total_out)
        waste_rate = 0.0 if total_out == 0 else round(waste_qty / total_out * 100, 1)

        return jsonify({
            "success": True,
            "data": {
                "total_restock_events": int(total_restock),
                "active_ingredients": int(active_items),
                "waste_rate_percent": waste_rate,
                "period": {
                    "start_date": start.isoformat(),
                    "end_date": end.isoformat()
                }
            }
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

@bp_owner_reports.route("/api/owner/reports/restock-frequency", methods=["GET"])
def restock_frequency():
    start, end = _get_range()
    end_inclusive = end + timedelta(days=1)
    conn = get_conn()
    cur = conn.cursor()
    try:
        # ===== CURRENT PERIOD =====
        cur.execute("""
            SELECT 
              COUNT(*) AS import_count,
              DATEDIFF(MAX(created_at), MIN(created_at)) AS span_days
            FROM transactions
            WHERE type='Import'
              AND created_at BETWEEN %s AND %s
        """, (start, end_inclusive))
        import_count, span_days = cur.fetchone()
        import_count = import_count or 0
        span_days = span_days or 0
        if import_count <= 1:
            avg_freq = span_days or 0
        else:
            avg_freq = round(span_days / (import_count - 1), 1)

        # ===== PREVIOUS PERIOD (để so sánh) =====
        # cùng độ dài với kỳ hiện tại
        period_delta = (end - start)
        prev_start = start - period_delta
        prev_end = start
        prev_end_inclusive = prev_end

        cur.execute("""
            SELECT 
              COUNT(*) AS import_count,
              DATEDIFF(MAX(created_at), MIN(created_at)) AS span_days
            FROM transactions
            WHERE type='Import'
              AND created_at BETWEEN %s AND %s
        """, (prev_start, prev_end_inclusive))
        prev_import_count, prev_span_days = cur.fetchone()
        prev_import_count = prev_import_count or 0
        prev_span_days = prev_span_days or 0
        if prev_import_count <= 1:
            prev_avg_freq = prev_span_days or 0
        else:
            prev_avg_freq = round(prev_span_days / (prev_import_count - 1), 1)

        # chênh lệch (dương = nhanh hơn, vì số ngày giữa 2 lần nhập giảm)
        avg_freq_diff = None
        avg_freq_trend = "same"
        if prev_avg_freq > 0:
            avg_freq_diff = round(prev_avg_freq - avg_freq, 1)
            if avg_freq_diff > 0:
                avg_freq_trend = "faster"   # nhanh hơn kỳ trước
            elif avg_freq_diff < 0:
                avg_freq_trend = "slower"   # chậm hơn
            else:
                avg_freq_trend = "same"

        prev_period_label = prev_start.strftime("%b")  # ví dụ 'Aug'

        # ===== Most frequent item =====
        cur.execute("""
            SELECT i.ingredient_id, i.name, COUNT(*) AS times
            FROM transactions t
            JOIN batches b ON t.batch_id = b.batch_id
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE t.type='Import'
              AND t.created_at BETWEEN %s AND %s
            GROUP BY i.ingredient_id
            ORDER BY times DESC
            LIMIT 1
        """, (start, end_inclusive))
        row = cur.fetchone()
        most_item = None
        if row:
            ing_id, ing_name, times = row
            most_item = {
                "ingredient_id": int(ing_id),
                "ingredient_name": ing_name,
                "restock_times": int(times or 0),
            }

        # ===== Supplier reliability (như cũ) =====
        cur.execute("""
            SELECT COUNT(DISTINCT b.batch_id)
            FROM transactions t
            JOIN batches b ON t.batch_id = b.batch_id
            WHERE t.type='Import'
              AND t.created_at BETWEEN %s AND %s
        """, (start, end_inclusive))
        total_batches = cur.fetchone()[0] or 0

        cur.execute("""
            SELECT COUNT(DISTINCT a.batch_id)
            FROM alerts a
            WHERE a.alert_type IN ('NearExpiry','Expired')
              AND a.created_at BETWEEN %s AND %s
        """, (start, end_inclusive))
        bad_batches = cur.fetchone()[0] or 0

        if total_batches == 0:
            reliability = 100.0
        else:
            reliability = round(max(0.0, 100.0 - bad_batches / total_batches * 100.0), 1)

        # ===== Detailed log (giống bản cũ) =====
        cur.execute("""
            SELECT
              i.ingredient_id,
              i.name AS ingredient_name,
              COUNT(*) AS freq,
              AVG(t.quantity) AS avg_qty,
              MAX(t.created_at) AS last_restock,
              MIN(t.created_at) AS first_restock
            FROM transactions t
            JOIN batches b ON t.batch_id = b.batch_id
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE t.type='Import'
              AND t.created_at BETWEEN %s AND %s
            GROUP BY i.ingredient_id
            ORDER BY freq DESC, ingredient_name
            LIMIT 50
        """, (start, end_inclusive))
        rows = cur.fetchall()

        # map tần suất kỳ trước theo ingredient
        cur.execute("""
            SELECT i.ingredient_id, COUNT(*) AS freq
            FROM transactions t
            JOIN batches b ON t.batch_id = b.batch_id
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE t.type='Import'
              AND t.created_at BETWEEN %s AND %s
            GROUP BY i.ingredient_id
        """, (prev_start, prev_end))
        prev_map = {rid: int(freq or 0) for (rid, freq) in cur.fetchall()}

        log = []
        for ing_id, name, freq, avg_qty, last_restock, first_restock in rows:
            freq = int(freq or 0)
            avg_qty = _to_float(avg_qty)

            if first_restock and last_restock and freq > 1:
                span = (last_restock - first_restock).days
                avg_gap = span / (freq - 1) if span > 0 else avg_freq
            else:
                avg_gap = avg_freq or 0

            next_expected = (last_restock.date() + timedelta(days=avg_gap)) if last_restock else None

            prev_freq = prev_map.get(ing_id, 0)
            if prev_freq == 0 and freq == 0:
                trend_str = "Stable"
            elif freq > prev_freq:
                trend_str = "Rising"
            elif freq < prev_freq:
                trend_str = "Falling"
            else:
                trend_str = "Stable"

            log.append({
                "ingredient_id": int(ing_id),
                "ingredient_name": name,
                "frequency": freq,
                "avg_quantity": round(avg_qty, 2),
                "last_restock": last_restock.isoformat() if last_restock else None,
                "next_expected": next_expected.isoformat() if next_expected else None,
                "reliability_percent": reliability,
                "trend": trend_str
            })

        # ===== Import trend 6 tháng =====
        cur.execute("""
            SELECT 
              DATE_FORMAT(t.created_at, '%Y-%m') AS ym,
              COUNT(*) AS import_events,
              COUNT(DISTINCT b.ingredient_id) AS ingredient_count,
              SUM(t.quantity) AS total_qty
            FROM transactions t
            JOIN batches b ON t.batch_id = b.batch_id
            WHERE t.type = 'Import'
            GROUP BY ym
            ORDER BY ym
        """)

        sql_map = {}
        all_months = []
        for ym, events, ing_count, total_qty in cur.fetchall():
            all_months.append(ym)
            sql_map[ym] = {
                "events": int(events or 0),
                "ingredient_count": int(ing_count or 0),
                "total_qty": _to_float(total_qty),
            }

        trend = []
        if all_months:
            all_months = sorted(set(all_months))
            last_6 = all_months[-6:]
            for ym in last_6:
                data = sql_map.get(ym, {
                    "events": 0,
                    "ingredient_count": 0,
                    "total_qty": 0.0,
                })
                trend.append({
                    "label": ym,
                    "imports": data["events"],
                    "ingredient_count": data["ingredient_count"],
                    "total_qty": round(data["total_qty"], 2),
                })
        else:
            # nếu chưa có Import nào thì hiển thị 6 tháng hiện tại = 0
            current_month = date.today().replace(day=1)
            for offset in range(5, -1, -1):
                m_date = _add_months(current_month, -offset)
                ym = m_date.strftime("%Y-%m")
                trend.append({
                    "label": ym,
                    "imports": 0,
                    "ingredient_count": 0,
                    "total_qty": 0.0,
                })

        return jsonify({
            "success": True,
            "data": {
                "summary": {
                    "avg_restock_frequency_days": avg_freq,
                    "prev_avg_restock_frequency_days": prev_avg_freq,
                    "avg_freq_diff_days": avg_freq_diff,
                    "avg_freq_trend": avg_freq_trend,      # faster / slower / same
                    "prev_period_label": prev_period_label,
                    "most_frequent_item": most_item,
                    "supplier_reliability_percent": reliability
                },
                "log": log,
                "trend": trend,
                "period": {
                    "start_date": start.isoformat(),
                    "end_date": end.isoformat()
                }
            }
        })

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()




def _add_months(d, months):
    """Cộng/trừ số tháng cho một date, trả về ngày 1 của tháng đó."""
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    return date(year, month, 1)
@bp_owner_reports.route("/api/owner/reports/inventory-analysis", methods=["GET"])
def inventory_analysis():
    start, end = _get_range()
    end_inclusive = end + timedelta(days=1)
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        # ----- Summary KPIs -----
        cur.execute("SELECT COUNT(*) AS total_items FROM ingredients")
        total_items = cur.fetchone()["total_items"] or 0

        cur.execute("""
            SELECT COUNT(*) AS active_items
            FROM inventory
            WHERE current_stock > 0
        """)
        active_items = cur.fetchone()["active_items"] or 0

        cur.execute("""
            SELECT COUNT(*) AS low_stock
            FROM inventory
            WHERE current_stock <= 0
        """)
        low_stock = cur.fetchone()["low_stock"] or 0

        cur.execute("""
            SELECT COUNT(DISTINCT batch_id) AS expired
            FROM batches
            WHERE status = 'Expired'
        """)
        expired = cur.fetchone()["expired"] or 0

        # Turnover: stock_now / usage_per_day
        cur.execute("""
            SELECT 
              SUM(CASE WHEN type IN ('Use','Export') THEN quantity ELSE 0 END) AS used_qty,
              SUM(CASE WHEN type='Waste' THEN quantity ELSE 0 END) AS waste_qty,
              SUM(CASE WHEN type IN ('Use','Waste','Export') THEN quantity ELSE 0 END) AS total_out
            FROM transactions
            WHERE created_at BETWEEN %s AND %s
        """, (start, end_inclusive))
        row = cur.fetchone()
        used_qty = _to_float(row["used_qty"])
        waste_qty = _to_float(row["waste_qty"])
        total_out = _to_float(row["total_out"])

        days = max((end - start).days, 1)
        daily_usage = used_qty / days if days else 0
        cur.execute("SELECT SUM(current_stock) AS stock_now FROM inventory")
        stock_now = _to_float(cur.fetchone()["stock_now"])
        avg_turnover_days = 0 if daily_usage == 0 else round(stock_now / daily_usage, 1)
        waste_rate = 0 if total_out == 0 else round(waste_qty / total_out * 100, 1)

                 # ----- Stock movement: lấy theo tất cả dữ liệu, rồi chọn 5 tháng gần nhất có data -----
        cur.execute("""
            SELECT 
            DATE_FORMAT(created_at, '%Y-%m') AS ym,
            SUM(CASE WHEN type = 'Import' THEN quantity ELSE 0 END)          AS incoming,
            SUM(CASE WHEN type IN ('Use','Export') THEN quantity ELSE 0 END) AS outgoing,
            SUM(CASE WHEN type = 'Waste' THEN quantity ELSE 0 END)           AS waste
            FROM transactions
            GROUP BY ym
            ORDER BY ym
        """)


        sql_map = {}
        for r in cur.fetchall():
            ym = r["ym"]
            sql_map[ym] = {
                "incoming": _to_float(r["incoming"]),
                "outgoing": _to_float(r["outgoing"]),
                "waste": _to_float(r["waste"]),
            }

        movement = []

        if sql_map:
            # Lấy 5 tháng cuối cùng có dữ liệu
            all_months = sorted(sql_map.keys())  # ['2025-09', '2025-10', ...]
            last_5 = all_months[-5:]
            for ym in last_5:
                data = sql_map[ym]
                movement.append({
                    "label": ym,
                    "incoming": round(data["incoming"], 2),
                    "outgoing": round(data["outgoing"], 2),
                    "waste": round(data["waste"], 2),
                })
        else:
            # Nếu không có transaction nào thì fallback: 5 tháng hiện tại = 0
            current_month = date.today().replace(day=1)
            for offset in range(4, -1, -1):
                m_date = _add_months(current_month, -offset)
                ym = m_date.strftime("%Y-%m")
                movement.append({
                    "label": ym,
                    "incoming": 0.0,
                    "outgoing": 0.0,
                    "waste": 0.0,
                })



                        # ----- Top used ingredients -----
        cur.execute("""
            SELECT i.name, SUM(t.quantity) AS qty
            FROM transactions t
            JOIN batches b ON t.batch_id = b.batch_id
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE t.type IN ('Use','Export')
              AND t.created_at BETWEEN %s AND %s
            GROUP BY i.ingredient_id
            ORDER BY qty DESC
            LIMIT 5
        """, (start, end_inclusive))

        rows = cur.fetchall()
        total_used_all = sum(_to_float(r["qty"]) for r in rows) or 1.0

        top_used = []
        max_qty = _to_float(rows[0]["qty"]) if rows else None
        for idx, r in enumerate(rows, start=1):
            qty = _to_float(r["qty"])
            pct_index = 0 if not max_qty else round(qty / max_qty * 100, 1)
            share = round(qty / total_used_all * 100, 1)
            top_used.append({
                "rank": idx,
                "name": r["name"],
                "usage_index": pct_index,
                "total_qty": round(qty, 2),
                "share_percent": share
            })




        # ----- Expiry risk alert -----
        cur.execute("""
            SELECT b.batch_id, b.lot_code, i.name, b.expiry_date
            FROM batches b
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE b.status = 'NearExpiry'
            ORDER BY b.expiry_date ASC
            LIMIT 1
        """)
        row = cur.fetchone()
        expiry_risk = None
        if row:
            days_left = (row["expiry_date"] - date.today()).days if row["expiry_date"] else None
            expiry_risk = {
                "ingredient_name": row["name"],
                "batch_code": row["lot_code"],
                "days_left": days_left
            }

        return jsonify({
            "success": True,
            "data": {
                "summary": {
                    "total_items": int(total_items),
                    "active": int(active_items),
                    "low_stock": int(low_stock),
                    "expired": int(expired),
                    "avg_turnover_days": avg_turnover_days,
                    "waste_rate_percent": waste_rate
                },
                "movement": movement,
                "top_used": top_used,
                "expiry_risk": expiry_risk
            }
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

@bp_owner_reports.route("/api/owner/reports/waste-summary", methods=["GET"])
def waste_summary():
    start, end = _get_range()
    end_inclusive = end + timedelta(days=1)
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        # ===== Tổng waste (kg) trong kỳ =====
        cur.execute("""
            SELECT quantity, unit
            FROM Waste_Reports
            WHERE report_date BETWEEN %s AND %s
        """, (start, end_inclusive))
        total_kg = 0.0
        for r in cur.fetchall():
            total_kg += _unit_to_kg(r["quantity"], r["unit"])

        # ===== Waste rate = Waste / (Use + Waste + Export) =====
        cur.execute("""
            SELECT
              SUM(CASE WHEN type='Waste' THEN quantity ELSE 0 END) AS waste_qty,
              SUM(CASE WHEN type IN ('Use','Waste','Export') THEN quantity ELSE 0 END) AS total_out
            FROM transactions
            WHERE created_at BETWEEN %s AND %s
        """, (start, end_inclusive))
        row = cur.fetchone()
        waste_rate = 0.0
        if row:
            w = _to_float(row["waste_qty"])
            t = _to_float(row["total_out"])
            waste_rate = 0 if t == 0 else round(w / t * 100, 1)

        # ===== Breakdown by reason (Expiration / Spoilage / Production Error / Other) =====
        cur.execute("""
            SELECT reason, quantity, unit
            FROM Waste_Reports
            WHERE report_date BETWEEN %s AND %s
        """, (start, end_inclusive))
        reason_map = {}
        for r in cur.fetchall():
            reason_raw = (r["reason"] or "").lower()
            if "expir" in reason_raw:
                cat = "Expiration"
            elif "spoil" in reason_raw or "storage" in reason_raw:
                cat = "Spoilage / Storage"
            elif "error" in reason_raw or "production" in reason_raw:
                cat = "Production Error"
            else:
                cat = "Other"
            reason_map.setdefault(cat, 0.0)
            reason_map[cat] += _unit_to_kg(r["quantity"], r["unit"])

        breakdown = []
        top_cause = None
        max_kg = 0.0
        for cat, kg in reason_map.items():
            pct = 0 if total_kg == 0 else round(kg / total_kg * 100, 1)
            breakdown.append({"reason": cat, "kg": round(kg, 2), "percent": pct})
            if kg > max_kg:
                max_kg = kg
                top_cause = cat

        # ===== Highest Waste "Categories" – dùng tên nguyên liệu làm category =====
        cur.execute("""
            SELECT i.name AS ingredient_name, wr.quantity, wr.unit
            FROM Waste_Reports wr
            JOIN batches b ON wr.batch_id = b.batch_id
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE wr.report_date BETWEEN %s AND %s
        """, (start, end_inclusive))
        cat_map = {}
        for r in cur.fetchall():
            cat = r["ingredient_name"] or "Uncategorized"
            cat_map.setdefault(cat, 0.0)
            cat_map[cat] += _unit_to_kg(r["quantity"], r["unit"])

        categories = sorted(
            [
                {"category": c, "kg": round(kg, 2)}
                for c, kg in cat_map.items()
            ],
            key=lambda x: x["kg"],
            reverse=True
        )

        # ===== Top wasted ingredients log (không dùng i.category nữa) =====
        cur.execute("""
            SELECT
              i.name AS ingredient_name,
              wr.quantity,
              wr.unit,
              wr.reason,
              wr.report_date,
              b.status AS batch_status
            FROM Waste_Reports wr
            JOIN batches b ON wr.batch_id = b.batch_id
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE wr.report_date BETWEEN %s AND %s
            ORDER BY wr.report_date DESC
            LIMIT 50
        """, (start, end_inclusive))
        top_wasted = []
        for r in cur.fetchall():
            reason_raw = (r["reason"] or "").lower()
            if "expir" in reason_raw:
                reason_type = "Expiration"
            elif "spoil" in reason_raw or "storage" in reason_raw:
                reason_type = "Spoilage / Storage"
            elif "error" in reason_raw or "production" in reason_raw:
                reason_type = "Production Error"
            else:
                reason_type = "Other"

            status_view = r.get("batch_status") or "Disposed"

            top_wasted.append({
                "ingredient_name": r["ingredient_name"],
                "category": r["ingredient_name"],      # dùng tên nguyên liệu như category hiển thị
                "quantity": _to_float(r["quantity"]),
                "unit": r["unit"],
                "reason": r["reason"] or "Unknown",
                "reason_type": reason_type,
                "report_date": r["report_date"].strftime("%Y-%m-%d"),
                "status": status_view
            })

        # ===== Critical alert: so sánh category có waste lớn nhất với kỳ trước =====
        critical_alert = None
        if categories:
            top_cat = categories[0]["category"]
            curr_kg = categories[0]["kg"]

            prev_start = start - (end - start)
            prev_end = start
            cur.execute("""
                SELECT i.name AS ingredient_name, wr.quantity, wr.unit
                FROM Waste_Reports wr
                JOIN batches b ON wr.batch_id = b.batch_id
                JOIN ingredients i ON b.ingredient_id = i.ingredient_id
                WHERE wr.report_date BETWEEN %s AND %s
            """, (prev_start, prev_end))
            prev_map = {}
            for r in cur.fetchall():
                cat = r["ingredient_name"] or "Uncategorized"
                prev_map.setdefault(cat, 0.0)
                prev_map[cat] += _unit_to_kg(r["quantity"], r["unit"])

            prev_kg = prev_map.get(top_cat, 0.0)
            if prev_kg > 0:
                change = round((curr_kg - prev_kg) / prev_kg * 100, 1)
                if change >= 10:
                    critical_alert = {
                        "category": top_cat,
                        "change_percent": change,
                        "message": f"Waste for {top_cat} increased by {change}% compared to previous period."
                    }

        rescued_kg = 0.0  # sau này nếu có logic rescue thì cập nhật sau

        return jsonify({
            "success": True,
            "data": {
                "summary": {
                    "total_waste_kg": round(total_kg, 2),
                    "waste_rate_percent": waste_rate,
                    "top_cause": top_cause,
                    "rescued_kg": round(rescued_kg, 2)
                },
                "breakdown": breakdown,
                "categories": categories,
                "top_wasted": top_wasted,
                "critical_alert": critical_alert,
                "period": {
                    "start_date": start.isoformat(),
                    "end_date": end.isoformat()
                }
            }
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()




@bp_owner_reports.route("/api/owner/reports/restock-frequency/export", methods=["GET"])
def export_restock_frequency():
    start, end = _get_range()
    end_inclusive = end + timedelta(days=1)

    conn = get_conn()
    cur = conn.cursor()
    try:
        # Lấy log giống phần trong restock_frequency
        cur.execute("""
            SELECT
              i.name AS ingredient_name,
              COUNT(*) AS freq,
              AVG(t.quantity) AS avg_qty,
              MAX(t.created_at) AS last_restock,
              MIN(t.created_at) AS first_restock
            FROM transactions t
            JOIN batches b ON t.batch_id = b.batch_id
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE t.type='Import'
              AND t.created_at BETWEEN %s AND %s
            GROUP BY i.ingredient_id
            ORDER BY freq DESC, ingredient_name
        """, (start, end_inclusive))
        rows = cur.fetchall()

        # Tạo file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "RestockFrequency"
        ws.append(["Ingredient", "Frequency", "Avg Quantity", "First Restock", "Last Restock"])

        for name, freq, avg_qty, last_restock, first_restock in rows:
            ws.append([
                name,
                int(freq or 0),
                float(avg_qty or 0),
                first_restock.strftime("%Y-%m-%d") if first_restock else "",
                last_restock.strftime("%Y-%m-%d") if last_restock else "",
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"restock_frequency_{start}_{end}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()
@bp_owner_reports.route("/api/owner/reports/inventory-analysis/export", methods=["GET"])
def export_inventory_analysis():
    start, end = _get_range()
    end_inclusive = end + timedelta(days=1)
    conn = get_conn()
    cur = conn.cursor()
    try:
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "StockMovement"
        ws1.append(["Year-Month", "Incoming", "Outgoing", "Waste"])

        cur.execute("""
            SELECT 
              DATE_FORMAT(created_at, '%Y-%m') AS ym,
              SUM(CASE WHEN type='Import' THEN quantity ELSE 0 END) AS incoming,
              SUM(CASE WHEN type IN ('Use','Export') THEN quantity ELSE 0 END) AS outgoing,
              SUM(CASE WHEN type='Waste' THEN quantity ELSE 0 END) AS waste
            FROM transactions
            WHERE created_at >= DATE_SUB(%s, INTERVAL 5 MONTH)
              AND created_at <= %s
            GROUP BY ym
            ORDER BY ym
        """, (end_inclusive, end_inclusive))

        for ym, incoming, outgoing, waste in cur.fetchall():
            ws1.append([
                ym,
                float(incoming or 0),
                float(outgoing or 0),
                float(waste or 0),
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"inventory_analysis_{start}_{end}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

@bp_owner_reports.route("/api/owner/reports/waste-summary/export", methods=["GET"])
def export_waste_summary_excel():
    start, end = _get_range()
    end_inclusive = end + timedelta(days=1)
    conn = get_conn()
    cur = conn.cursor()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "WasteLog"
        ws.append(["Date", "Ingredient", "Quantity", "Unit", "Reason"])

        cur.execute("""
            SELECT
              wr.report_date,
              i.name AS ingredient_name,
              wr.quantity,
              wr.unit,
              wr.reason
            FROM Waste_Reports wr
            JOIN batches b ON wr.batch_id = b.batch_id
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE wr.report_date BETWEEN %s AND %s
            ORDER BY wr.report_date DESC
        """, (start, end_inclusive))
        for report_date, ingredient_name, qty, unit, reason in cur.fetchall():
            ws.append([
                report_date.strftime("%Y-%m-%d"),
                ingredient_name,
                float(qty or 0),
                unit,
                reason,
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"waste_summary_{start}_{end}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

@bp_owner_reports.route("/api/owner/reports/production-summary", methods=["GET"])
def production_summary():
    """
    Tạm hiểu Production = lượng nguyên liệu được Use + Waste trong kỳ.
    """
    start, end = _get_range()
    end_inclusive = end + timedelta(days=1)
    conn = get_conn()
    cur = conn.cursor(dictionary=True)
    try:
        # Summary
        cur.execute("""
            SELECT
              SUM(CASE WHEN type IN ('Use','Export') THEN quantity ELSE 0 END) AS used_qty,
              SUM(CASE WHEN type='Waste' THEN quantity ELSE 0 END) AS waste_qty,
              COUNT(DISTINCT DATE(created_at)) AS active_days
            FROM transactions
            WHERE created_at BETWEEN %s AND %s
        """, (start, end_inclusive))
        row = cur.fetchone()
        used_qty = _to_float(row["used_qty"])
        waste_qty = _to_float(row["waste_qty"])
        active_days = row["active_days"] or 0

        total_production = used_qty + waste_qty
        avg_daily_usage = 0 if active_days == 0 else round(used_qty / active_days, 2)
        production_waste_rate = 0 if total_production == 0 else round(waste_qty / total_production * 100, 1)

        # Daily line
        cur.execute("""
            SELECT
              DATE(created_at) AS d,
              SUM(CASE WHEN type IN ('Use','Export') THEN quantity ELSE 0 END) AS used_qty,
              SUM(CASE WHEN type='Waste' THEN quantity ELSE 0 END) AS waste_qty
            FROM transactions
            WHERE created_at BETWEEN %s AND %s
            GROUP BY DATE(created_at)
            ORDER BY d
        """, (start, end_inclusive))
        daily = []
        for r in cur.fetchall():
            daily.append({
                "date": r["d"].strftime("%Y-%m-%d"),
                "used_qty": _to_float(r["used_qty"]),
                "waste_qty": _to_float(r["waste_qty"])
            })

        # Top used ingredients (giống inventory_analysis)
        cur.execute("""
            SELECT i.name, SUM(t.quantity) AS qty
            FROM transactions t
            JOIN batches b ON t.batch_id = b.batch_id
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE t.type IN ('Use','Export')
              AND t.created_at BETWEEN %s AND %s
            GROUP BY i.ingredient_id
            ORDER BY qty DESC
            LIMIT 5
        """, (start, end_inclusive))
        top_used = []
        rows = cur.fetchall()
        max_qty = _to_float(rows[0]["qty"]) if rows else None
        for idx, r in enumerate(rows, start=1):
            pct = 0 if not max_qty else round(_to_float(r["qty"]) / max_qty * 100, 1)
            top_used.append({
                "rank": idx,
                "name": r["name"],
                "usage_index": pct
            })

        return jsonify({
            "success": True,
            "data": {
                "summary": {
                    "total_used_qty": round(used_qty, 2),
                    "total_waste_qty": round(waste_qty, 2),
                    "avg_daily_usage": avg_daily_usage,
                    "production_waste_rate": production_waste_rate
                },
                "daily": daily,
                "top_used": top_used,
                "period": {
                    "start_date": start.isoformat(),
                    "end_date": end.isoformat()
                }
            }
        })
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()
@bp_owner_reports.route("/api/owner/reports/production-summary/export", methods=["GET"])
def export_production_summary():
    start, end = _get_range()
    end_inclusive = end + timedelta(days=1)
    conn = get_conn()
    cur = conn.cursor()
    try:
        wb = Workbook()

        # Sheet 1: Daily production
        ws1 = wb.active
        ws1.title = "DailyProduction"
        ws1.append(["Date", "UsedQty", "WasteQty"])

        cur.execute("""
            SELECT
              DATE(created_at) AS d,
              SUM(CASE WHEN type IN ('Use','Export') THEN quantity ELSE 0 END) AS used_qty,
              SUM(CASE WHEN type='Waste' THEN quantity ELSE 0 END) AS waste_qty
            FROM transactions
            WHERE created_at BETWEEN %s AND %s
            GROUP BY DATE(created_at)
            ORDER BY d
        """, (start, end_inclusive))
        for d, used_qty, waste_qty in cur.fetchall():
            ws1.append([
                d.strftime("%Y-%m-%d"),
                float(used_qty or 0),
                float(waste_qty or 0),
            ])

        # Sheet 2: top used ingredients
        ws2 = wb.create_sheet(title="TopUsedIngredients")
        ws2.append(["Rank", "Ingredient", "UsageIndex"])

        cur.execute("""
            SELECT i.name, SUM(t.quantity) AS qty
            FROM transactions t
            JOIN batches b ON t.batch_id = b.batch_id
            JOIN ingredients i ON b.ingredient_id = i.ingredient_id
            WHERE t.type='Use'
              AND t.created_at BETWEEN %s AND %s
            GROUP BY i.ingredient_id
            ORDER BY qty DESC
            LIMIT 5
        """, (start, end_inclusive))
        rows = cur.fetchall()
        max_qty = rows[0][1] if rows else None
        for idx, (name, qty) in enumerate(rows, start=1):
            pct = 0 if not max_qty else round(float(qty or 0) / float(max_qty or 1) * 100, 1)
            ws2.append([idx, name, pct])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"production_summary_{start}_{end}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

